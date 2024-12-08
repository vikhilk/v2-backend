

from datetime import datetime, timedelta
from xmlrpc.client import DateTime
from flask import Flask,request,jsonify,session
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
import os
import json
from flask_cors import CORS, cross_origin
import pandas as pd
# from tablib import Dateset
from sqlalchemy import desc
import random
# from sqlalchemy.sql import test
from sqlalchemy import ForeignKey,select,desc
from sqlalchemy.sql.expression import func
from exchangelib import Credentials, Account, Configuration, Version, Build, DELEGATE, NTLM,Message, Mailbox,FileAttachment
from exchangelib.protocol import BaseProtocol, NoVerifyHTTPAdapter
import urllib3
from markupsafe import Markup
from sqlalchemy.exc import IntegrityError

from cryptography.fernet import Fernet
import openpyxl 
from openpyxl.styles import PatternFill, Alignment
from openpyxl.comments import Comment
import hashlib



#Init app
app = Flask(__name__)
mail = Mail(app)
app.app_context().push()
cors = CORS(app)
app.config['CORS_HEADERS'] = 'application/json'
basedir = os.path.abspath(os.path.dirname(__file__))

#Datebase
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir,'db.sqlite')
app.config['SQLACHEMY_TRACK_MODIFICATION'] = False




#Init db
db = SQLAlchemy(app)

#Init ma
ma = Marshmallow(app)


alpha=['a','b','c','d','e','f','g','h','i','j']
roman=['i','ii','iii','iv','v','vi','vii','viii','ix','x']
#Quize Class/Model
class Quize(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(200))
    pair1 = db.Column(db.String(200))
    pair2 = db.Column(db.String(200))
    img_link = db.Column(db.String(500))
    que = db.Column(db.String(200),unique=True)
    opt1 = db.Column(db.String(200))
    opt2 = db.Column(db.String(200))
    opt3 = db.Column(db.String(200))
    opt4 = db.Column(db.String(200))
    correct_ans = db.Column(db.String(200))
    subcategory = db.Column(db.String(200))
    category_id = db.Column(ForeignKey('category.id'))
    explanation = db.Column(db.String(500))
    level = db.Column(db.String(200))
    adminID=db.Column(db.String(200))

    db.create_all()

    def __init__(self,type,pair1,pair2,img_link,que,opt1,opt2,opt3,opt4,correct_ans,subcategory,category_id,explanation,level,adminID):
        self.type=type
        self.pair1=pair1
        self.pair2=pair2
        self.img_link=img_link
        self.que=que
        self.opt1= opt1 
        self.opt2= opt2
        self.opt3= opt3
        self.opt4= opt4
        self.correct_ans= correct_ans
        self.subcategory=subcategory
        self.category_id=category_id
        self.explanation=explanation
        self.level=level
        self.adminID=adminID

#Quize schema
class QuizeSchema(ma.Schema):
    class Meta:
        fields = ('id','type','pair1','pair2','img_link','que','opt1','opt2','opt3','opt4','correct_ans','subcategory','category_id','explanation','level','adminID')

#Init Schema
Quize_schema = QuizeSchema()
Quizes_schema = QuizeSchema(many=True)


class Subcategory(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    topic_name=db.Column(db.String(200))
    category_id=db.Column(ForeignKey('category.id'))

    admin_id=db.Column(db.String(200))

def __init__(self,topic_name,category_id,admin_id):

    self.topic_name=topic_name
    self.category_id=category_id
    self.admin_id=admin_id

class subcategorySchema(ma.Schema):
    class Meta:
        fields=('id','topic_name','category_id','admin_id')

subcategory_schema= subcategorySchema()
subcategorys_schema= subcategorySchema(many=True)


#category
class Category(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    Category_name=db.Column(db.String(200),unique=True)
    admin_id=db.Column(db.String(200))
    category_Statuses = db.relationship('CategoryStatus',backref='category',lazy=True)

    def add_admin_ids(self,admin_id):
        if not self.admin_id:
            self.admin_id=str(admin_id)
        else:
            self.admin_id += f",{admin_id}"


    def __init__(self, Category_name,admin_id):
        self.Category_name = Category_name
        self.admin_id= str(admin_id)


#category Schema
class CategorySchema(ma.Schema):
    class Meta:
        fields = ('id', 'Category_name','admin_id')

#init schema
category_schema = CategorySchema()
categorys_schema = CategorySchema(many=True)



#Login Model
class Login(db.Model):

    id = db.Column(db.Integer, primary_key=True,autoincrement=True)
    name = db.Column(db.String(200))
    username = db.Column(db.String(200))
    password = db.Column(db.String(200))
    email = db.Column(db.String(200))
    status = db.Column(db.Boolean,default=True,nullable=False)
    BU=db.Column(ForeignKey('admin_login.BU'))
    mode=db.Column(db.String(200),default=True,nullable=False)


    def __init__(self,name,username,password,email,status,BU,mode):
      
        self.name = name
        self.username = username
        self.set_password(password)
        self.email = email
        self.status = status
        self.BU = BU
        self.mode = mode

    # update user status
    def update_user_status(self,status):
        self.status = status

    # update user exam mode e.g. practice/certification
    def update_user_mode(self,mode):
        self.mode = mode

    def set_password(self,password):
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        self.password = hashed_password

    def check_password(self,password):
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        return self.password == hashed_password
# Quize schema
class LoginSchema(ma.Schema):
    class Meta:
        fields = ('id','name','username','password','email','status','BU','mode')

#Init Schema
login_schema = LoginSchema()
logins_schema = LoginSchema(many=True) 


# admin login 

class adminLogin(db.Model):

    id = db.Column(db.String(200), primary_key=True)
    name = db.Column(db.String(200))
    username = db.Column(db.String(200))
    password = db.Column(db.String(200))
    email = db.Column(db.String(200))
    BU = db.Column(db.String(200))
    noOfquestion=db.Column(db.String(200))
    timerrPerTest=db.Column(db.String(200))


    def __init__(self,id,name,username,password,email,BU,noOfquestion,timerrPerTest):
        self.id=id
        self.name = name
        self.username = username
        self.set_password(password)
        self.email = email
        self.BU = BU
        self.noOfquestion = noOfquestion
        self.timerrPerTest = timerrPerTest
        

    def set_password(self,password):
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        self.password = hashed_password

    def check_password(self,password):
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        return self.password == hashed_password
# Quize schema
class adminLoginSchema(ma.Schema):
    class Meta:
        fields = ('id','name','username','password','email','BU','noOfquestion','timerrPerTest')

#Init Schema
adminLogin_schema = adminLoginSchema()
adminLogins_schema = adminLoginSchema(many=True) 


# Subcategory Status Model
class SubcategoryStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    status = db.Column(db.Boolean)
    user_id = db.Column(db.String(200))
    topic_name = db.Column(db.String(200))
    level = db.Column(db.String(200))
    category_id = db.Column(db.String(200))

    def __init__(self, status, user_id, topic_name, level, category_id):
        self.status = status
        self.user_id = user_id
        self.topic_name = topic_name
        self.level = level
        self.category_id = category_id

class SubcategoryStatusSchema(ma.Schema):
    class Meta:
        fields = ('id','status','user_id','topic_name','level','category_id')

subcategoryStatus_schema = SubcategoryStatusSchema()
subcategorysStatus_schema = SubcategoryStatusSchema(many=True) 



# Category Status Model
class CategoryStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    status = db.Column(db.Boolean)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    user_id = db.Column(db.String(200))

    def __init__(self, status, category_id, user_id):
        self.status = status
        self.category_id = category_id
        self.user_id = user_id

class CategoryStatusSchema(ma.Schema):
    class Meta:
        fields = ('id','status','category_id','user_id')

CategoryStatus_schema = CategoryStatusSchema()
CategorysStatus_schema = CategoryStatusSchema(many=True) 




#AdminDashboard Model
class AdminDashboard(db.Model):
    id=db.Column(db.Integer, primary_key=True)
    login_id = db.Column(db.Integer)
    marksGot = db.Column(db.Integer)
    AvgTime = db.Column(db.Integer)
    correct_ans = db.Column(db.String(200))
    que_attempt = db.Column(db.String(200))
    email = db.Column(db.String(200))
    category_id = db.Column(ForeignKey('category.id'))

    db.create_all()

    def __init__(self,login_id,marksGot,AvgTime,correct_ans,que_attempt,email,category_id):
        self.login_id = login_id
        self.marksGot = marksGot
        self.AvgTime = AvgTime
        self.correct_ans = correct_ans
        self.que_attempt = que_attempt
        self.email = email
        self.category_id = category_id

# Quize schema
class AdminDashboardSchema(ma.Schema):
    class Meta:
        fields = ('id','login_id','AvgTime','marksGot','correct_ans','que_attempt','email','category_id')

#Init Schema
admindashboard_schema = AdminDashboardSchema()
admindashboards_schema = AdminDashboardSchema(many=True)


#chart
class UserChart(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    Subject = db.Column(db.String(200))
    attempt = db.Column(db.Integer)
    # created_date = db.Column(db.DateTime, default=db.func.current_timestamp())
    login_id = db.Column(ForeignKey('login.id'))



    db.create_all()

    def __init__(self,Subject,attempt,login_id):
        self.Subject = Subject
        self.attempt = attempt
        self.login_id = login_id

class UserChartSchema(ma.Schema):
    class Meta:
        fields = ('id','Subject','attempt','login_id')

userchart_schema = UserChartSchema()
usercharts_schema = UserChartSchema(many=True)



# admin login API for loginto portal
@app.route('/login_admin',methods=['POST'])


def adminlogin():
    data=request.json
    id=data.get('id')
    password=data.get('password')

    hashed_password=hashlib.sha256(password.encode()).hexdigest()
    user=adminLogin.query.filter_by(id=id,password=hashed_password).first()

    if user:
        return adminLogin_schema.jsonify(user)
    else:
        return "error"


# create admin user 
@app.route('/add_admin', methods=['POST'])

def add_admin():
    try:
        # Parse the incoming JSON data
        data = request.get_json()

        # Validate required fields
        required_fields = ['id', 'name', 'username', 'password', 'email', 'BU', 'noOfquestion', 'timerrPerTest']
        missing_fields = [field for field in required_fields if field not in data]
        if missing_fields:
            return jsonify({'error': f'Missing required fields: {", ".join(missing_fields)}'}), 400

        # Create a new adminLogin instance
        new_admin = adminLogin(
            id=data['id'],
            name=data['name'],
            username=data['username'],
            password=data['password'],  # Password is hashed in set_password()
            email=data['email'],
            BU=data['BU'],
            noOfquestion=data['noOfquestion'],
            timerrPerTest=data['timerrPerTest']
        )

        # Add the new admin to the database and commit
        db.session.add(new_admin)
        db.session.commit()

        return jsonify({'message': 'Admin added successfully!'}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Get all category quiz
@app.route('/quizes_all/<category_id>', methods=['GET'])
@cross_origin()
def get_allCategpry_quiz(category_id):
    categories = db.session.query(Quize.category_id).distinct()
    selected_quizes = []

    for category_tuple in categories:
        category_id = category_tuple[0]
        category_quizes = Quize.query.filter_by(category_id=str(category_id)).order_by(Quize.id).limit(5).all()
        selected_quizes.extend(category_quizes)

    result = Quizes_schema.dump(selected_quizes)
    return jsonify(result)

# Update the exam mode i.e. Certificate or Practice
@app.route('/userMode/<id>', methods=['PUT'])
@cross_origin()
def update_user_mode(id):
    try:
        mode = request.json['new_mode']
        user = Login.query.get(id)

        if user is None:
            return jsonify({"message": "mode not found"}), 404

        user.update_user_mode(mode)
        db.session.commit()

        serialized_user = login_schema.dump(user)
        return jsonify({"message": "user mode updated successfully", "user": serialized_user}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Enable/disable user
@app.route('/users/<id>', methods=['PUT'])
@cross_origin()
def update_user_status(id):
    try:
        status = request.json['new_status']
        user = Login.query.get(id)

        if user is None:
            return jsonify({"message": "Category not found"}), 404

        user.update_user_status(status)
        db.session.commit()

        serialized_user = login_schema.dump(user)
        return jsonify({"message": "user status updated successfully", "user": serialized_user}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Add category status
@app.route('/AddCategoryStatus', methods=['POST'])
@cross_origin()
def add_category_status():
    data = request.json
    category_id = data.get('category_id')
    user_id = data.get('user_id')
    new_status = data.get('status')

    existing_status = CategoryStatus.query.filter_by(category_id=category_id, user_id=user_id).first()
    if existing_status:
        if existing_status.status == new_status:
            return jsonify({'message': 'status is the same'})
        else:
            existing_status.status = new_status
            db.session.commit()
            return jsonify({'message': 'status updated successfully'})
    else:
        new_category_status = CategoryStatus(status=new_status, category_id=category_id, user_id=user_id)
        db.session.add(new_category_status)
        db.session.commit()
        return jsonify({'message': 'categorystatus created'})


# Get category-wise subcategory topic names for user list
@app.route('/GetTopicLevels/<admin_id>/<category_id>/<string:topic_name>', methods=['GET'])
@cross_origin()
def getTopicLevel(admin_id, category_id, topic_name):
    adminidsplited = admin_id.split(',')
    results = db.session.query(Quize.level)\
        .filter(Quize.adminID.in_(adminidsplited))\
        .filter(Quize.category_id == category_id)\
        .filter(Quize.subcategory == topic_name).distinct()

    levels = [result[0] for result in results]
    return levels, 200

# Get subcategory by admin ID and category ID
@app.route('/get_subcategory/<admin_id>/<category_id>', methods=['GET'])
@cross_origin()
def get_subcategories(admin_id, category_id):
    subcategories = Subcategory.query.filter_by(admin_id=admin_id, category_id=category_id).all()
    if not subcategories:
        return jsonify({"message": "No subcategories found for given admin_id and category_id"}), 404

    result = subcategorys_schema.dump(subcategories)
    return jsonify(result), 200

# Enable/disable category
@app.route('/category/<int:category_id>', methods=['PUT'])
@cross_origin()
def update_category_status(category_id):
    try:
        status = request.json['new_status']
        category = Category.query.get(category_id)
        if category is None:
            return jsonify({"message": "Category not found"}), 404

        category.update_status(status)
        db.session.commit()

        serialized_category = category_schema.dump(category)
        return jsonify({"message": "Category status updated successfully", "category": serialized_category}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Update the number of questions and time per test
@app.route('/admin/update/<int:admin_id>', methods=['PUT'])
@cross_origin()
def update_admin(admin_id):
    try:
        admin_login = AdminLogin.query.get(admin_id)
        if admin_login is None:
            return jsonify({"message": "Admin not found"}), 404

        admin_login.noOfquestion = request.json['noOfquestion']
        admin_login.timerPerTest = request.json['timerPerTest']
        db.session.commit()

        serialized_admin = adminLogin_schema.dump(admin_login)
        return jsonify({"message": "Admin updated successfully", "admin": serialized_admin}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Get the category names for user list by admin ID
@app.route('/get_category_by_admin_id/<admin_id>', methods=['GET'])
@cross_origin()
def get_category_by_admin_id(admin_id):
    categories = Category.query.filter(Category.admin_id.contains(admin_id)).all()
    if categories:
        category_names = [{"id": category.id, "category_name": category.Category_name} for category in categories]
        return jsonify(category_names)
    else:
        return jsonify([]), 404

# Add category name and status in category table
@app.route('/category', methods=['POST'])
@cross_origin()
def add_cat():
    Category_name = request.json['Category_name']
    new_category = Category(Category_name=Category_name, status=True)
    db.session.add(new_category)
    db.session.commit()
    db.create_all()
    return category_schema.jsonify(new_category)

# Get quiz by user id and category id 
@app.route('/quiz/<user_id>/<category_id>', methods=['GET'])
@cross_origin()
def get_filtered_quizes(user_id, category_id):
    try:
        subcategory_statuses = SubcategoryStatus.query.filter_by(
            user_id=user_id).all()

        formatted_quizes = []

        for subcategory_status in subcategory_statuses:
            topic_name = subcategory_status.topic_name
            level = subcategory_status.level
            status = subcategory_status.status

            if status:
                quizes = Quize.query.filter_by(
                    subcategory=topic_name, level=level, category_id=category_id).all()

                for quize in quizes:
                    formatted_quiz = {
                        "type": quize.type,
                        "pair1": quize.pair1,
                        "pair2": quize.pair2,
                        "img_link": quize.img_link,
                        "que": quize.que,
                        "opt1": quize.opt1,
                        "opt2": quize.opt2,
                        "opt3": quize.opt3,
                        "opt4": quize.opt4,
                        "correct_ans": quize.correct_ans,
                        "subcategory": quize.subcategory,
                        "category_id": quize.category_id,
                        "explanation": quize.explanation,
                        "level": quize.level,
                        "adminID": quize.adminID
                    }
                    formatted_quizes.append(formatted_quiz)

        return jsonify(formatted_quizes)
    except Exception as e:
        return jsonify({"error": str(e)}), 500



#create a quize
@app.route('/quize',methods=['POST'])
@cross_origin()
def add_quize():
    type=request.json['type']
    pair1=request.json['pair1']
    pair2=request.json['pair2']
    img_link=request.json['img_link']
    que=request.json['que']
    opt1=request.json['opt1']
    opt2=request.json['opt2']
    opt3=request.json['opt3']
    opt4=request.json['opt4']
    correct_ans=request.json['correct_ans']
    category_id=request.json['category_id']
    explanation=request.json['explanation']
    level=request.json['level']


    new_quize = Quize(type,pair1,pair2,img_link,que,opt1,opt2,opt3,opt4,correct_ans,category_id,explanation,level)
    db.session.add(new_quize)
    db.session.commit()
    db.create_all()

    return Quize_schema.jsonify(new_quize)




def type_validation(ques_json):
    if(ques_json["type"]=="MCQ" or ques_json["type"]=="MPF" or ques_json["type"]=="IMG"):
        return True, "Valid Question"
    else:
        return False, "Invalid or blank type"
    

def pair_validation(ques_json):
    if (ques_json["type"]=="MPF"):
        if (ques_json["pair1"]== None or ques_json["pair2"]==None):
            return False, "pair1 and/or pair2 column is blank"
        else:
            if(ques_json["pair1"].count("\n")==ques_json["pair2"].count("\n")):
                return True, "Valid Question"
            else:
                return False, "pair1 and pair2 do not have the same number of options"
    else:
        return True, "Valid Question"
    
def options_validation (ques_json):
    if (ques_json["opt1"]== None or ques_json["opt2"]==None or ques_json["opt3"]== None or ques_json["opt4"]==None):
        return False, "One or more of the options are blank"
    else:
        if (str(ques_json["opt1"]).strip()==str(ques_json["opt2"]).strip() or str(ques_json["opt1"]).strip()==str(ques_json["opt3"]).strip() or str(ques_json["opt1"]).strip()==str(ques_json["opt4"]).strip() or str(ques_json["opt2"]).strip()==str(ques_json["opt3"]).strip() or str(ques_json["opt2"]).strip()==str(ques_json["opt4"]).strip() or str(ques_json["opt3"]).strip()==str(ques_json["opt4"]).strip() ):
            return False, "Duplicate options"
        else:
            return True, "Valid Question"

def explanation_validation (ques_json):
    if (ques_json["explanation"] == None):
        return False, "explanation is blank"
    else:
        return True, "Valid Question"

def subcategory_validation(ques_json):
    if (ques_json["subcategory"] == None):
        return False, "subcategory is blank"
    else:
        return True, "Valid Question"


def correct_ans_validation (ques_json):
    if (ques_json["correct_ans"] == None):
        return False, "correct_ans is blank"
    else:
        if (str(ques_json["correct_ans"]).strip()==str(ques_json["opt1"]).strip() or                                                                                   str(ques_json["correct_ans"]).strip()==str(ques_json["opt2"]).strip() or                                                                                       str(ques_json["correct_ans"]).strip()==str(ques_json["opt3"]).strip() or                                                                                        str(ques_json["correct_ans"]).strip()==str(ques_json["opt4"]).strip()): 
            return True, "Valid Question"
        else:
            return False, "correct_ans does not match any of the options"


def level_validation (ques_json):
    if (ques_json["level"] == None):
        return False, "level is blank"
    else:
        if(ques_json["level"]=="High" or ques_json["level"]=="Low"):
            return True, "Valid Question"
        else:
            return False, "Invalid level"


def IMG_validation(ques_json):
    if (ques_json["type"]=="IMG"):
        if (ques_json["img_link"] == None):
            return False, "img_link is blank"
        else:
            return True, "Valid Question"
    else:
        return True, "Valid Question"


def pair_validation_MCQ_IMG(ques_json):

    if (ques_json["type"] == "MCQ" or ques_json["type"] == "IMG"):

        if (ques_json["pair1"] == None and ques_json["pair2"] == None):

            return True, "Valid Question"

        else:

            return False, "pair1/pair2 should be blank for IMG and MCQ type questions"

    else:

        return True, "Valid Question"

def img_link_validation_MCQ_MPF (ques_json):



    if (ques_json["type"] == "MCQ" or ques_json["type"] == "MPF"):

        if (ques_json["img_link"] == None):



            return True, "Valid Question"



        else:



            return False, "img_link should be blank for MPF and MCQ type questions"

    else:

        return True, "Valid Question"




# excel upload  

# Excel upload route
# @app.route('/UploadExcel', methods=["POST"])
# @cross_origin()
# def submitExcelFile():
#     subcategories = {}   # Use a set to store unique subcategories from the Excel file
#     duplicate_data = []
#     quizzes = []

#     data = request.files['file']
#     adminID = json.loads(request.form['json'])
#     adminID = str(adminID)
#     print(adminID)

#     data.save(data.filename)
#     df_excel = pd.ExcelFile(data.filename)
#     wb_obj = openpyxl.load_workbook(data.filename)
#     redFill=PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
    
#     list_of_dfs = []
#     validation_success = True
    
#     try:
#         for sheet in df_excel.sheet_names:
#             print(sheet)
#             if sheet == 'Sheet2':
#                 continue
            
#             # Check if the category already exists
#             existing_category = Category.query.filter(Category.Category_name == sheet).first()
#             if not existing_category:
#                 new_category = Category(Category_name=sheet, admin_id=adminID)
#                 db.session.add(new_category)
#                 db.session.commit()
#                 db.create_all()
#                 catID = new_category.id
#                 print(catID)
#             else:
#                 existing_admin_ids = existing_category.admin_id.split(',')
#                 if adminID not in existing_admin_ids:
#                     existing_admin_ids.append(adminID)
#                     existing_category.admin_id = ",".join(existing_admin_ids)
#                     db.session.commit()
#                 catID = existing_category.id

#             category_sheet = wb_obj[sheet]
#             max_row = 0
#             question_set_for_category = []

#             # For loop to build valid dataframe and highlight the invalid data
#             category_sheet.cell(row=1, column=14).value = "Comment"
#             i = 2

#             while max_row == 0:
#                 ques_json = {
#                     "type": category_sheet.cell(row=i, column=1).value,
#                     "pair1": category_sheet.cell(row=i, column=2).value,
#                     "pair2": category_sheet.cell(row=i, column=3).value,
#                     "img_link": category_sheet.cell(row=i, column=4).value,
#                     "que": category_sheet.cell(row=i, column=5).value,
#                     "opt1": category_sheet.cell(row=i, column=6).value,
#                     "opt2": category_sheet.cell(row=i, column=7).value,
#                     "opt3": category_sheet.cell(row=i, column=8).value,
#                     "opt4": category_sheet.cell(row=i, column=9).value,
#                     "correct_ans": category_sheet.cell(row=i, column=10).value,
#                     "explanation": category_sheet.cell(row=i, column=11).value,
#                     "level": category_sheet.cell(row=i, column=12).value,
#                     "subcategory": category_sheet.cell(row=i, column=13).value,
#                 }
#                 if all(val is None for val in ques_json.values()):
#                     break
#                 else:
#                     comment = ""
#                     type_validation_status, type_validation_comment = type_validation(ques_json)
#                     if not type_validation_status:
#                         comment += type_validation_comment

#                     pair_validation_status, pair_validation_comment = pair_validation(ques_json)
#                     if not pair_validation_status:
#                         comment += pair_validation_comment

#                     options_validation_status, options_validation_comment = options_validation(ques_json)
#                     if not options_validation_status:
#                         comment += options_validation_comment

#                     explanation_validation_status, explanation_validation_comment = explanation_validation(ques_json)
#                     if not explanation_validation_status:
#                         comment += explanation_validation_comment
                    
#                     correct_ans_validation_status, correct_ans_validation_comment = correct_ans_validation(ques_json)
#                     if (not correct_ans_validation_status):
#                         comment = comment+";"+ correct_ans_validation_comment 

#                     level_validation_status, level_validation_comment = level_validation(ques_json)
#                     if (not level_validation_status):
#                         comment = comment+";"+ level_validation_comment  


#                     IMG_validation_status, IMG_validation_comment = IMG_validation(ques_json)
#                     if (not IMG_validation_status):
#                         comment = comment+";"+ IMG_validation_comment


#                     img_link_validation_MCQ_MPF_status, img_link_validation_MCQ_MPF_comment = img_link_validation_MCQ_MPF(ques_json)
#                     if (not img_link_validation_MCQ_MPF_status):
#                         comment = comment+";"+ img_link_validation_MCQ_MPF_comment  

#                     pair_validation_MCQ_IMG_status, pair_validation_MCQ_IMG_comment = pair_validation_MCQ_IMG(
#                         ques_json)
#                     if (not pair_validation_MCQ_IMG_status):
#                         comment = comment+";"+ pair_validation_MCQ_IMG_comment  

                    
#                     if (comment == ""):
#                         comment = "Valid Question"
#                         question_set_for_category.append(ques_json)
#                         category_sheet.cell(row = i, column = 14).value = "Valid Question"
#                     else:
#                         validation_success = False

#                         category_sheet.cell(row = i, column = 13).value = comment[1:]

#                         for k in range(1,14):
#                             category_sheet.cell(row=i, column=k).fill = redFill
#                     i+=1
        
#             df1=pd.DataFrame.from_dict(question_set_for_category)
#             df1['category_id']=catID

#             list_of_dfs.append(df1)

#             for index,row in df1.iterrows():
#                 subcategory = row['subcategory']

#                 if subcategory not in subcategories:
#                     subcategories[subcategory] = {'category_ids': []}
#                 subcategories[subcategory]['category_ids'].append(catID)

#         for subcategory, info in subcategories.items():
#             for category_id in info['category_ids']:
#                 existing_subcategory= Subcategory.query.filter_by(topic_name=subcategory,category_id=category_id,admin_id=adminID).first()
#                 if not existing_subcategory:
#                     new_subcategory = Subcategory(
#                         topic_name=subcategory,
#                         category_id=category_id,
#                         admin_id=adminID
#                     )
#                     db.session.add(new_subcategory)
#         db.session.commit()

#         df= pd.concat(list_of_dfs,ignore_index=True)
#         print(df)

#         wb_obj.save("Validation Results.xlsx")

#         matched_df = df
#         def convrt_to_html(pair1,pair2):
#             table_html = "Match the following<br><table style='border:1px solid black;'>" 
#             table_html += "<tr><th style='border:1px solid black;'>Column A</th><th style='border:1px solid black;'>Column B</th></tr>" 
#             pair1_values = pair1.split('\n')
#             pair2_values = pair2.split('\n')

#             i=0

#             for p1,p2 in zip(pair1_values,pair2_values):
#                 table_html += "<tr><td style='border:1px solid black;'>" +alpha[i]+" {}</td><td style='border:1px solid black;'>".format(p1) +roman[i]+" {}</td></tr>".format(p2)
#                 i+=1
#             table_html += "</table>"
#             return table_html
#         matched_df['que']=df.apply(lambda row: convrt_to_html(row['pair1'],row['pair2'])if row['type']=='MPF'else row['que'],axis=1)

#         def convert_html_img_tag(img_link,que):
#             html_img_tag ='<img src="' + img_link + '"><br><p>'+ que +'</p>'
#             return Markup(html_img_tag)
#         matched_df['que']=df.apply(lambda row: convert_html_img_tag(row['img_link'],row['que'])if row['type']=='IMG'else row['que'],axis=1)


#         df_json = json.loads(matched_df.to_json(orient="records"))


#         for json_data in df_json:
#             type_str=str(json_data['type']).strip()
#             pair1_str=str(json_data['pair1']).strip() if isinstance(json_data['pair1'],str) else None
#             pair2_str=str(json_data['pair2']).strip() if isinstance(json_data['pair2'],str) else None
#             img_link_str=str(json_data['img_link']).strip()
#             que_str=str(json_data['que']).strip()
#             opt1_str=str(json_data['opt1']).strip()
#             opt2_str=str(json_data['opt2']).strip()
#             opt3_str=str(json_data['opt3']).strip()
#             opt4_str=str(json_data['opt4']).strip()
#             correct_ans_str=str(json_data['correct_ans']).strip()
#             subcategory_str=str(json_data['subcategory']).strip()
#             explanation_str=str(json_data['explanation']).strip()
#             category_id_str=str(json_data['category_id']).strip()
#             level_str=str(json_data['level']).strip()

#             existing_quiz=Quize.query.filter_by(que=que_str,category_id=category_id_str,subcategory=subcategory_str).first()

#             if not existing_quiz:
#                 new_quize=Quize(
#                     type=type_str,
#                     pair1=pair1_str,
#                     pair2=pair2_str,
#                     img_link=img_link_str,
#                     que=que_str,
#                     opt1=opt1_str,
#                     opt2=opt2_str,
#                     opt3=opt3_str,
#                     opt4=opt4_str,
#                     correct_ans=correct_ans_str,
#                     explanation=explanation_str,
#                     level=level_str,
#                     subcategory=subcategory_str,
#                     category_id=category_id_str,
#                     adminID=adminID
#                 )
#                 db.session.add(new_quize)

#             else:

#                 existing_admin_ids=existing_quiz.adminID.split(',')
#                 if adminID not in existing_admin_ids:
#                     existing_admin_ids.append(adminID)
#                     existing_quiz.adminID=','.join(existing_admin_ids)
#             db.session.commit()
#     except IntegrityError as e:
#         db.session.rollback()
#         validation_success=False
    
#     if validation_success:
#         msg={'msg':'success'}
#         return jsonify(msg)
#     else:
       
#        msg= {"msg": "Failed",  "error": "Validation failed for some questions"}
  
#        return jsonify(msg)



                    

    






# excel upload data into db
# @app.route('/UploadExcel', methods=["POST"])
# @cross_origin()
# def submitExcelFile():
#     duplicate_data=[]
#     quezes = []
#     data = request.files['file']
#     data.save(data.filename)
#     df_excel = pd.ExcelFile(data.filename)
#     wb_obj = openpyxl.load_workbook(data.filename)
#     redFill = PatternFill(start_color='FFFF0000',
#                           end_color='FFFF0000',
#                           fill_type='solid')
    
#     # df=pd.read_excel(data)
#     #convert pandas to json
#     list_of_dfs =[]

#     #Iterate trough each worksheet

#     validation_success = True
#     for sheet in df_excel.sheet_names:
#         print(sheet)
#         if (sheet=='sheet2'):
#             continue

#         cat = Category.query.filter(Category.Category_name == sheet).all()
#         result = categorys_schema.dump(cat)
#         if len(result)==0:
#             new_category = Category(Category_name=sheet, status=True)
#             db.session.add(new_category)
#             db.session.commit()
#             db.create_all()
#             catId= new_category.id

#         else:
#             catId=result[0]["id"]



#         category_sheet = wb_obj[sheet]
#         max_row = 0
#         question_set_for_category = []
#         #for loop to build valid dataframe and highlight the invalid data
#         category_sheet.cell(row = 1, column = 13).value = "comment"
#         i=2
#         while(max_row==0):
#             ques_json = {
#                 "type": category_sheet.cell(row=i, column=1).value,
#                 "pair1": category_sheet.cell(row=i, column=2).value,
#                 "pair2": category_sheet.cell(row=i, column=3).value,
#                 "img_link": category_sheet.cell(row=i, column=4).value,
#                 "que": category_sheet.cell(row=i, column=5).value,
#                 "opt1": category_sheet.cell(row=i, column=6).value,
#                 "opt2": category_sheet.cell(row=i, column=7).value,
#                 "opt3": category_sheet.cell(row=i, column=8).value,
#                 "opt4": category_sheet.cell(row=i, column=9).value,
#                 "correct_ans": category_sheet.cell(row=i, column=10).value,
#                 "explanation": category_sheet.cell(row=i, column=11).value,
#                 "level": category_sheet.cell(row=i, column=12).value,
#             }
#             print("=====================================================")
#             print(ques_json)
#             if ques_json['type'] == None and ques_json['pair1']==None and ques_json['pair2']==None and ques_json['que']==None and ques_json['img_link']==None and ques_json['opt1']==None and ques_json['opt2']==None and ques_json['opt3']==None and ques_json['opt4']==None and ques_json['correct_ans']==None and ques_json['explanation']==None and ques_json['level']==None:
#                 print("inside if statement")
#                 break
#             else:
#                 print("inside else statement")
#                 comment =""

#                 type_validation_status,type_validation_comment = type_validation(ques_json)
#                 if (not type_validation_status):
#                     comment = comment+";"+ type_validation_comment  

#                 pair_validation_status,pair_validation_comment = pair_validation(
#                     ques_json)
#                 if (not pair_validation_status):
#                     comment = comment+";"+ pair_validation_comment

#                 options_validation_status,options_validation_comment = options_validation(ques_json)
#                 if (not options_validation_status):
#                     comment = comment+";"+ options_validation_comment    

#                 explanation_validation_status,explanation_validation_comment = explanation_validation(ques_json)
#                 if (not explanation_validation_status):
#                     comment = comment+";"+ explanation_validation_comment 

#                 correct_ans_validation_status, correct_ans_validation_comment = correct_ans_validation(ques_json)
#                 if (not correct_ans_validation_status):
#                     comment = comment+";"+ correct_ans_validation_comment 

#                 level_validation_status, level_validation_comment = level_validation(ques_json)
#                 if (not level_validation_status):
#                     comment = comment+";"+ level_validation_comment  


#                 IMG_validation_status, IMG_validation_comment = IMG_validation(ques_json)
#                 if (not IMG_validation_status):
#                     comment = comment+";"+ IMG_validation_comment


#                 img_link_validation_MCQ_MPF_status, img_link_validation_MCQ_MPF_comment = img_link_validation_MCQ_MPF(ques_json)
#                 if (not img_link_validation_MCQ_MPF_status):
#                     comment = comment+";"+ img_link_validation_MCQ_MPF_comment  

#                 pair_validation_MCQ_IMG_status, pair_validation_MCQ_IMG_comment = pair_validation_MCQ_IMG(
#                     ques_json)
#                 if (not pair_validation_MCQ_IMG_status):
#                     comment = comment+";"+ pair_validation_MCQ_IMG_comment  

                
#                 if (comment == ""):
#                     comment = "Valid Question"
#                     question_set_for_category.append(ques_json)
#                     category_sheet.cell(row = i, column = 13).value = "Valid Question"
#                 else:
#                     validation_success = False

#                     category_sheet.cell(row = i, column = 13).value = comment[1:]

#                     for k in range(1,14):
#                         category_sheet.cell(row=i, column=k).fill = redFill
#             i+=1



#         #df1 = df_excel.parse(sheet)
#         print(question_set_for_category)
#         df1 = pd.DataFrame.from_dict(question_set_for_category)
#         df1['category_id'] = catId
#         #And append it to the list
#         list_of_dfs.append(df1)

#     #combine all DataFrame into one
#     df = pd.concat(list_of_dfs,ignore_index=True)
#     wb_obj.save("Validation Results.xlsm")


#     matched_df = df
#     def convrt_to_html(pair1,pair2):
#         table_html = "Match the following<br><table style='border:1px solid black;'>" 
#         table_html += "<tr><th style='border:1px solid black;'>Column A</th><th style='border:1px solid black;'>Column B</th></tr>" 
#         pair1_values = pair1.split('\n')
#         pair2_values = pair2.split('\n')

#         i=0

#         for p1,p2 in zip(pair1_values,pair2_values):
#             table_html += "<tr><td style='border:1px solid black;'>" +alpha[i]+" {}</td><td style='border:1px solid black;'>".format(p1) +roman[i]+" {}</td></tr>".format(p2)
#             i+=1
#         table_html += "</table>"
#         return table_html
#     matched_df['que']=df.apply(lambda row: convrt_to_html(row['pair1'],row['pair2'])if row['type']=='MPF'else row['que'],axis=1)

#     def convert_html_img_tag(img_link,que):
#         html_img_tag ='<img src="' + img_link + '"><br><p>'+ que +'</p>'
#         return Markup(html_img_tag)
#     matched_df['que']=df.apply(lambda row: convert_html_img_tag(row['img_link'],row['que'])if row['type']=='IMG'else row['que'],axis=1)


#     df_json = json.loads(matched_df.to_json(orient="records"))















#     print(df_json)
#     for jsondata in df_json:
#         jsondata['type']=str(jsondata['type']).strip()
#         jsondata['pair1']=str(jsondata['pair1']).strip() if isinstance(jsondata['pair1'],str) else None
#         jsondata['pair2']=str(jsondata['pair2']).strip() if isinstance(jsondata['pair2'],str) else None
#         jsondata['que']=str(jsondata['que']).strip()
#         jsondata['opt1']=str(jsondata['opt1']).strip()
#         jsondata['opt2']=str(jsondata['opt2']).strip()
#         jsondata['opt3']=str(jsondata['opt3']).strip()
#         jsondata['opt4']=str(jsondata['opt4']).strip()
#         jsondata['correct_ans']=str(jsondata['correct_ans']).strip()

#         jsondata['explanation']=str(jsondata['explanation']).strip()
#         jsondata['level']=str(jsondata['level']).strip()





#         new_quize = Quize(**jsondata)

#         quizes.append(new_quize)
#     print('******************************************************')
#     print(df_json)




#     for quiz_data in quizes:
#         try:

#             db.session.add(quiz_data)
#             db.session.commit()

#         except Exception as e:

#             db.session.rollback()
#             duplicate_data.append(quiz_data)
#             pass






#     # [db.session.refresh(quizedata) for quizedata in quizes]

#     if validation_success:
#         msg={'msg':'success'}
#         return jsonify(msg)
#     else:
#         msg={'msg': r'C:\Users\VK56833\OneDrive - Citi\Python Assignment\code\flask'}
#         return jsonify(msg)
    



#Get all Quizes
@app.route('/quizes',methods=['Get'])
@cross_origin()
def get_quizes():
    all_quizes = Quize.query.all()
    all_quizes= Quize.query.order_by(func.random()).order_by(Quize.id).all()
    result = Quizes_schema.dump(all_quizes)


    return jsonify(result)
@app.route('/quizes_all/<category_id>',methods=['GET'])
@cross_origin()
def get_allCat_quiz(category_id):
    categories=db.session.query(Quize.category_id).distinct()
    selected_quizes=[]

    for category_tuple in categories:
        category_id=category_tuple[0]
        category_quizes=Quize.query.filter_by(category_id=str(category_id).order_by(Quize.id).limit(5).all())
        selected_quizes.extend(category_quizes)

    result = Quizes_schema.dump(selected_quizes)
    return jsonify(result)











# @app.route('/category/<int:category_id>',methods=['PUT'])
# def update_category_status(category_id):
#     try:
#         status = request.json['new_status']

#         category = Category.query.get(category_id)

#         if category is None:
#             return jsonify({"message": "Category not found"}), 404
        
#         category.update_status(status)
#         db.session.commit()

#         serialized_category = category_schema.dump(category)
#         return jsonify({"message": "Category status updated successfully","category":serialized_category}), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


@app.route('/category',methods=['POST'])
@cross_origin()

def add_categ():
    category_name = request.json['Category_name']

    new_category = Category(Category_name=Category_name,status=True)
    db.session.add(new_category)
    db.session.commit()
    db.create_all()

    return category_schema.jsonify(new_category)

@app.route('/category',methods=['GET'])
@cross_origin()
def get_category():
    all_category = Category.query.all()
    result = categorys_schema.dump(all_category)
    return jsonify(result)

#   all_quizes = Quize.query.order_by(func.random()).order_by(Quize.id).all()
#retrive the que by category id
@app.route('/category/<category_id>',methods=['GET'])
@cross_origin()
def get_categoryById(category_id):
    user = Quize.query.filter(Quize.category_id == category_id).order_by(func.random()).all()

    return Quizes_schema.jsonify(user)

@app.route('/catById/<id>',methods=['GET'])
@cross_origin()
def get_ByCategory(id):
    user = Category.query.get(id)
    return category_schema.jsonify(user)




#Create a user
@app.route('/user', methods=['POST'])
@cross_origin()
def add_user():
   
    name = request.json['name']
    username = request.json['username']
    password = request.json['password']
    email = request.json['email']
    status = request.json['status']
    BU = request.json['BU']
    mode = request.json['mode']

    # encrypted_password = hashlib.sha256(password.encode()).hexdigest()

    new_user = Login(name=name, username=username, 
                     password=password, email=email, status=status, BU=BU, mode=mode)
    db.session.add(new_user)
    db.session.commit()

    admin_ids = adminLogin.query.filter_by(BU=BU).with_entities(adminLogin.id).all()
    print(admin_ids)

    if admin_ids:
        admin_ids = [admin_id for admin_id, in admin_ids]
        associated_categories = Category.query.filter(Category.admin_id.contains(',' .join(map(str, admin_ids)))).all()

        unique_category_ids = set(category.id for category in associated_categories)
        print(unique_category_ids)
        for category_id in unique_category_ids:
            category_status = CategoryStatus(status=True, category_id=category_id, user_id=new_user.id)
            db.session.add(category_status)

        recipient = new_user.email
        cc_recipient = username
        subject = ' Catalyst Knowledge Assessment(CKA)'
        body = "Hi, \n\nThank you for registering in  Catalyst Knowledge Assessment. Please use your  email id and password to login into the portal. \n\nRegards, \nCatalyst Knowledge Assessment(CKA)"

        # send_email(account, subject, body, recipient, cc_recipient)
        db.session.commit()

        admin_record = adminLogin.query.filter_by(BU=BU).first()
        if admin_record:
            admin_id = admin_record.id
            subcategories = Subcategory.query.filter_by(admin_id=admin_id).all()

            for subcategory in subcategories:
                subcategory_status = SubcategoryStatus(
                    status=True,
                    category_id=subcategory.category_id,
                    topic_name=subcategory.topic_name,
                    user_id=new_user.id,
                    level="Low"
                )
                db.session.add(subcategory_status)
            db.session.commit()
        return login_schema.jsonify(new_user)

    return jsonify({"error": "Associated adminIds not found for the given BU"}), 400


#Login API
@app.route('/login',methods=['POST'])
@cross_origin()
def login():
    #if not request.is_json:
    #    return jsonify({"error": "Invalid request format"}), 400
    data=request.json
    email = data.get('email')
    password= data.get('password')

    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    user = Login.query.filter_by(email=email,password=hashed_password).first()

    if user:

        return login_schema.jsonify(user)
    else:
        return "error"
    
# get categoru names for userdashboard
@app.route('/GetCategoryName/<string:bu_name>/<string:user_id>', methods=['GET'])
def get_category_name(bu_name, user_id):
    # Query the Category and CategoryStatus tables based on admin_id and user_id
    results = db.session.query(Category, CategoryStatus, Login)\
        .join(CategoryStatus, Category.id == CategoryStatus.category_id)\
        .join(Login, Login.id == user_id)\
        .filter(Login.BU == bu_name)\
        .filter(CategoryStatus.user_id == user_id).all()

    if results:
        response_data = []
        for category, status, login in results:
            response_data.append({
                'category_name': category.Category_name,
                'status': status.status,
                'id': category.id
            })
        return jsonify(response_data)
    else:
        return jsonify({'message': 'Category not found'})



# Add category status
@app.route('/AddCategoryStatus', methods=['POST'])
def addCategoryStatus():
    data = request.json
    category_id = data.get('category_id')
    user_id = data.get('user_id')
    new_status = data.get('status')

    existing_status = CategoryStatus.query.filter_by(category_id=category_id, user_id=user_id).first()

    if existing_status:
        if existing_status.status != new_status:
            existing_status.status = new_status
            db.session.commit()
            return jsonify({"message": "status updated successfully"})
        else:
            return jsonify({"message": "status is the same"})
    else:
        new_category_status = CategoryStatus(status=new_status, category_id=category_id, user_id=user_id)
        db.session.add(new_category_status)
        db.session.commit()
        return jsonify({"message": "category status created"})

# Get category-wise subcategory topic names for user list
@app.route("/GetTopicLevels/<admin_id>/<category_id>/<string:topic_name>", methods=['GET'])
def getTopicLevels(admin_id, category_id, topic_name):
    adminidsplited = admin_id.split('.')
    results = db.session.query(Quize.level).filter(
        Quize.adminID.in_(adminidsplited),
        Quize.category_id == category_id,
        Quize.subcategory == topic_name
    ).distinct()

    levels = [result[0] for result in results]
    return levels, 200

# # Get subcategory by admin id and category id
# @app.route('/get_subcategory/<admin_id>/<category_id>', methods=['GET'])
# def get_subcategories(admin_id, category_id):
#     subcategories = Subcategory.query.filter_by(admin_id=admin_id, category_id=category_id).all()

#     if not subcategories:
#         return jsonify({"message": "No subcategories found for given admin_id and category_id"}), 404

#     result = subcategory_schema.dump(subcategories)
#     return jsonify(result), 200

# Enable/disable category

# @app.route('/category/<int:category_id>', methods=['PUT'])
# def update_category_status(category_id):
#     try:
#         status = request.json['new_status']
#         category = Category.query.get(category_id)

#         if category is None:
#             return jsonify({"message": "Category not found"}), 404

#         category.update_status(status)
#         db.session.commit()

#         serialized_category = category_schema.dump(category)
#         return jsonify({"message": "Category status updated successfully", "category": serialized_category}), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500





#get all users
@app.route('/users',methods=['GET'])
@cross_origin()
def get_users():
    all_users = Login.query.all()
    result = logins_schema.dump(all_users)
    return jsonify(result)

#get email of user


@app.route('/user_emailid/<login_id>', methods=['GET'])
@cross_origin()
def get_user_email(login_id):
    user_mail = Login.query.get(login_id)
    result = login_schema.dump(user_mail)

    return jsonify({'emailid': result['email']})
# get single user

def get_user(id):
    user = Login.query.get(id)
    return login_schema.jsonify(user)




@app.route('/userchart',methods=['POST'])
@cross_origin()

def userchart_data():

    Subject = request.json['Subject']
    attempt = request.json['attempt']
    login_id = request.json['login_id']


    # chart_data = UserChart(Subject, attempt,login_id)
    found_subj = UserChart.query.filter_by(Subject=Subject).first()
    if found_subj:
        #request.json['Subject'] == found_subj.Subject
        found_subj.attempt += 1
        db.session.commit()
        return userchart_schema.jsonify(found_subj)
    #else
    user = UserChart(Subject,attempt,login_id)

    db.session.add(user)
    db.session.commit()
    print(user)
    db.create_all()
    db.session.refresh(user)

    return userchart_schema.jsonify(user)

@app.route('/userchart/<login_id>',methods=['GET'])
@cross_origin()
def get_chartById(login_id):
    user = UserChart.query.filter(UserChart.login_id == login_id).all()
    print(user)

    return usercharts_schema.jsonify(user)







AdminDashboard.category = db.relationship(Category,primaryjoin=AdminDashboard.category_id==Category.id)
#Create a usser
@app.route('/admindata',methods=['POST'])
@cross_origin()

def admin_data():
    login_id = request.json['login_id']
    marksGot = request.json['marksGot']
    AvgTime = request.json['AvgTime']
    correct_ans = request.json['correct_ans']
    que_attempt = request.json['que_attempt']
    email = request.json['email']
    category_id = request.json['category_id']
    totalQue = request.json['totalQue']
    score = request.json['score']

    admin_table = AdminDashboard(login_id,marksGot,AvgTime,correct_ans,que_attempt,email,category_id)
    db.session.add(admin_table)
    db.session.commit()
    db.create_all()

    recipient=admin_table.email
    firstname= Login.query.filter(Login.id==login_id).all()[0].name
    cc_recipient = Login.query.filter(Login.id==login_id).all()[0].username
    categoryname= Category.query.filter(Category.id==category_id).all()[0].Category_name
    subject="AKA -" + categoryname + "Test Results"
    body = "Hi " + firstname+ ",\nThanks for taking the "+ categoryname +" test. Please find below a summary of the result: \n\tTotal Questions: " + str(totalQue) + "\n\tAttempted Questions: " + str(que_attempt)+"\n\tCorrect Answers: " + str(correct_ans)+"\n\tScore: " +str(marksGot)+"%("+str(score)+")\nRegards,\nAutomation Knowledge Assessment(AKA)"

    
    return admindashboard_schema.jsonify(admin_table)

#get all users
@app.route('/admindata',methods=['GET'])
@cross_origin()
def get_admin_tbl():
    all_users_data = AdminDashboard.query.order_by(AdminDashboard.marksGot.desc())
    result = admindashboards_schema.dump(all_users_data)

    return jsonify(result)

@app.route('/admindata/<category_id>',methods=['GET'])
@cross_origin()
def get_admin_tblById(category_id):
    user = AdminDashboard.query.filter(AdminDashboard.category_id == category_id).all()
    print(user)


    return admindashboards_schema.jsonify(user)


@app.route('/admindatajoin/<login_id>',methods=['GET'])
@cross_origin()
def get_admin_data(login_id):
    results = db.session.query(AdminDashboard,Category).join(Category,AdminDashboard.category_id==Category.id).filter(AdminDashboard.login_id == login_id).all()

    data=[]

    for admin,category in results:
        data.append({
            'login_id':admin.login_id,
            'id':admin.id,
            'category_id':admin.category_id,
            'marksGot':admin.marksGot,
            'AvgTime':admin.AvgTime,
            'correct_ans':admin.correct_ans,
            'que_attempt':admin.que_attempt,
            'email':admin.email,
            'category_id':admin.category_id,
            'Category_name':category.Category_name            
        })
    return jsonify(data)





import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Gmail SMTP configuration
GMAIL_SMTP_SERVER = "smtp.gmail.com"
GMAIL_SMTP_PORT = 587

# Update these details with your Gmail credentials
app_email_address = "vikhilkhobragade@gmail.com"
password = "nfgu lesl miey udou"  # App Password if 2FA is enabled

def send_email(subject, body, recipient, cc_recipient=None):
    # Set up the MIME structure
    msg = MIMEMultipart()
    msg['From'] = app_email_address
    msg['To'] = recipient
    msg['Subject'] = subject
    if cc_recipient:
        msg['Cc'] = cc_recipient

    # Attach the email body
    msg.attach(MIMEText(body, 'plain'))

    # Recipients list
    recipients = [recipient]
    if cc_recipient:
        recipients.append(cc_recipient)

    # Set up the server and send the email
    try:
        server = smtplib.SMTP(GMAIL_SMTP_SERVER, GMAIL_SMTP_PORT)
        server.starttls()  # Start TLS for security
        server.login(app_email_address, password)  # Login to Gmail
        server.sendmail(app_email_address, recipients, msg.as_string())  # Send email
        server.quit()
        print("Email sent successfully.")
    except Exception as e:
        print(f"Error sending email: {e}")

# Usage
send_email(
    subject="Test Email",
    body="This is a test email sent from a Gmail account.",
    recipient="vikhilkhobragade@gmail.com",
    cc_recipient="vikhilkhobragade@gmail.com"
)


# @app.route("/sendmail/<login_id>")
#  @cross_origin()
# def mailbody(login_id):
#    user = Login.query.get(login_id)
#    email = user.email
#    subject="Citi Exam"
#    body="Hi ,\n\nThank you for giving an citi exam. \n\nThanks & Regards,\nCiti Exam"
#    m= Message(
#        account=account,
#        subject=subject,
#        body=body,
#        to_recipients = [email]
#    )
#    m.send()


#    return "success"












# @app.route("/UploadExcel", methods=["POST"])
# @cross_origin()
# def submitExcelFile():
#     # Use a set to store unique subcategories from the Excel file
#     subcategories = {}
#     subcategory_levels = {}
#     duplicate_data = []
#     quizzes = []

#     data = request.files['file']
#     adminID = json.loads(request.form['json'])
#     adminID = str(adminID)
#     print(adminID)

#     data.save(data.filename)

#     df_excel = pd.ExcelFile(data.filename)
#     wb_obj = openpyxl.load_workbook(data.filename)

#     redFill = PatternFill(start_color="FFFF0000",
#                           end_color="FFFF0000",
#                           fill_type="solid")

#     list_of_dfs = []
#     validation_success = True

#     try:
#         for sheet in df_excel.sheet_names:
#             if sheet == "Sheet2":
#                 continue

#             # Check if the category already exists
#             existing_category = Category.query.filter(
#                 Category.Category_name == sheet).first()

#             if not existing_category:
#                 new_category = Category(Category_name=sheet, admin_id=adminID)
#                 db.session.add(new_category)
#                 db.session.commit()
#                 db.create_all()
#                 catId = new_category.id
#                 print(catId)
#             else:
#                 existing_admin_ids = existing_category.admin_id.split(',')
#                 if adminID not in existing_admin_ids:
#                     existing_admin_ids.append(adminID)
#                     existing_category.admin_id = ','.join(existing_admin_ids)
#                     db.session.commit()
#                 catId = existing_category.id

#             category_sheet = wb_obj[sheet]
#             max_row = 0
#             question_set_for_category = []
#             category_sheet.cell(row=1, column=14).value = "Comment"
#             i = 2

#             while max_row == 0:
#                 ques_json = {
#                     "type": category_sheet.cell(row=i, column=1).value,
#                     "pair1": category_sheet.cell(row=i, column=2).value,
#                     "pair2": category_sheet.cell(row=i, column=3).value,
#                     "img_link": category_sheet.cell(row=i, column=4).value,
#                     "que": category_sheet.cell(row=i, column=5).value,
#                     "opt1": category_sheet.cell(row=i, column=6).value,
#                     "opt2": category_sheet.cell(row=i, column=7).value,
#                     "opt3": category_sheet.cell(row=i, column=8).value,
#                     "opt4": category_sheet.cell(row=i, column=9).value,
#                     "correct_ans": category_sheet.cell(row=i, column=10).value,
#                     "explanation": category_sheet.cell(row=i, column=11).value,
#                     "level": category_sheet.cell(row=i, column=12).value,
#                     "subcategory": category_sheet.cell(row=i, column=13).value,
#                 }

#                 if all(val is None for val in ques_json.values()):
#                     break
#                 else:
#                     comment = ""

#                     type_validation_status, type_validation_comment = type_validation(ques_json)
#                     if not type_validation_status:
#                         comment += ";" + type_validation_comment

#                     pair_validation_status, pair_validation_comment = pair_validation(ques_json)
#                     if not pair_validation_status:
#                         comment += ";" + pair_validation_comment

#                     options_validation_status, options_validation_comment = options_validation(ques_json)
#                     if not options_validation_status:
#                         comment += ";" + options_validation_comment

#                     explanation_validation_status, explanation_validation_comment = explanation_validation(ques_json)
#                     if not explanation_validation_status:
#                         comment += ";" + explanation_validation_comment

#                     subcategory_name_validation_status, subcategory_validation_comment = subcategory_validation(ques_json)
#                     if not subcategory_name_validation_status:
#                         comment += ";" + subcategory_validation_comment

#                     correct_ans_validation_status, correct_ans_validation_comment = correct_ans_validation(ques_json)
#                     if not correct_ans_validation_status:
#                         comment += ";" + correct_ans_validation_comment

#                     level_validation_status, level_validation_comment = level_validation(ques_json)
#                     if not level_validation_status:
#                         comment += ";" + level_validation_comment

#                     IMG_validation_status, IMG_validation_comment = IMG_validation(ques_json)
#                     if not IMG_validation_status:
#                         comment += ";" + IMG_validation_comment

#                     img_link_validation_MCQ_MPF_status, img_link_validation_MCQ_MPF_comment = img_link_validation_MCQ_MPF(ques_json)
#                     if not img_link_validation_MCQ_MPF_status:
#                         comment += ";" + img_link_validation_MCQ_MPF_comment

#                     pair_validation_MCQ_IMG_status, pair_validation_MCQ_IMG_comment = pair_validation_MCQ_IMG(ques_json)
#                     if not pair_validation_MCQ_IMG_status:
#                         comment += ";" + pair_validation_MCQ_IMG_comment

#                     if comment == "":
#                         comment = "Valid Question"
#                         question_set_for_category.append(ques_json)
#                         category_sheet.cell(row=i, column=14).value = "Valid Question"
#                     else:
#                         validation_success = False
#                         category_sheet.cell(row=i, column=14).value = comment[1:]
#                         for k in range(1, 14):
#                             category_sheet.cell(row=i, column=k).fill = redFill
#                     i += 1

#             df1 = pd.DataFrame.from_dict(question_set_for_category)
#             df1['category_id'] = catId
#             list_of_dfs.append(df1)

#         for index, row in df1.iterrows():
#             subcategory = row['subcategory']
#             if subcategory not in subcategories:
#                 subcategories[subcategory] = {'category_ids': [catId]}

#         for subcategory, info in subcategories.items():
#             for category_id in info['category_ids']:
#                 existing_subcategory = Subcategory.query.filter_by(
#                     topic_name=subcategory, category_id=category_id, admin_id=adminID).first()
#                 if not existing_subcategory:
#                     new_subcategory = Subcategory(
#                         topic_name=subcategory,
#                         category_id=category_id,
#                         admin_id=adminID
#                     )
#                     db.session.add(new_subcategory)
#                     print("subcategory added")
#         db.session.commit()

#         df = pd.concat(list_of_dfs, ignore_index=True)
#         wb_obj.save("Validation Results.xlsx")

#         matched_df = df

#         df_json = json.loads(matched_df.to_json(orient="records"))
#         print(df_json)

#         for json_data in df_json:
#             type_str = str(json_data['type']).strip()
#             pair1_str = str(json_data['pair1']).strip() if isinstance(json_data['pair1'], str) else None
#             pair2_str = str(json_data['pair2']).strip() if isinstance(json_data['pair2'], str) else None
#             img_link_str = str(json_data['img_link']).strip()
#             que_str = str(json_data['que']).strip()
#             opt1_str = str(json_data['opt1']).strip()
#             opt2_str = str(json_data['opt2']).strip()
#             opt3_str = str(json_data['opt3']).strip()
#             opt4_str = str(json_data['opt4']).strip()
#             correct_ans_str = str(json_data['correct_ans']).strip()
#             subcategory_str = str(json_data['subcategory']).strip()
#             explanation_str = str(json_data['explanation']).strip()
#             category_id_str = str(json_data['category_id']).strip()
#             level_str = str(json_data['level']).strip()

#             existing_quiz = Quize.query.filter_by(
#                 que=que_str, category_id=category_id_str, subcategory=subcategory_str).first()

#             if not existing_quiz:
#                 new_quiz = Quize(
#                     type=type_str,
#                     pair1=pair1_str,
#                     pair2=pair2_str,
#                     img_link=img_link_str,
#                     que=que_str,
#                     opt1=opt1_str,
#                     opt2=opt2_str,
#                     opt3=opt3_str,
#                     opt4=opt4_str,
#                     correct_ans=correct_ans_str,
#                     subcategory=subcategory_str,
#                     explanation=explanation_str,
#                     level=level_str,
#                     category_id=category_id_str,
#                     adminID=adminID
#                 )
#                 db.session.add(new_quiz)
#             else:
#                 existing_admin_ids = existing_quiz.adminID.split(',')
#                 if adminID not in existing_admin_ids:
#                     existing_admin_ids.append(adminID)
#                     existing_quiz.adminID = ','.join(existing_admin_ids)

#         db.session.commit()
#     except IntegrityError as e:
#         db.session.rollback()
#         validation_success = False

#     if validation_success:
#         msg = {'msg': 'success'}
#     else:
#         msg = {'msg': 'failed'}

#     return jsonify(msg)





@app.route("/UploadExcel", methods=["POST"])
@cross_origin()
def submitExcelFile():
    # Initialize variables
    subcategories = {}
    validation_success = True
    list_of_dfs = []

    # Retrieve uploaded file and admin ID
    data = request.files['file']
    adminID = json.loads(request.form['json'])
    adminID = str(adminID)

    # Save and load the Excel file
    data.save(data.filename)
    df_excel = pd.ExcelFile(data.filename)
    wb_obj = openpyxl.load_workbook(data.filename)

    # Red fill for invalid rows
    redFill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    try:
        # Process each sheet in the uploaded Excel file
        for sheet in df_excel.sheet_names:
            if sheet == "Sheet2":  # Skip specific sheets if necessary
                continue

            # Check or create the category
            existing_category = Category.query.filter(Category.Category_name == sheet).first()
            if not existing_category:
                new_category = Category(Category_name=sheet, admin_id=adminID)
                db.session.add(new_category)
                db.session.commit()
                catId = new_category.id
            else:
                # Update admin_id if necessary
                existing_admin_ids = existing_category.admin_id.split(',')
                if adminID not in existing_admin_ids:
                    existing_admin_ids.append(adminID)
                    existing_category.admin_id = ','.join(existing_admin_ids)
                    db.session.commit()
                catId = existing_category.id

            # Load the category sheet
            category_sheet = wb_obj[sheet]
            category_sheet.cell(row=1, column=14).value = "Comment"
            i = 2  # Start from the second row (assume first row is header)
            question_set_for_category = []

            while True:
                # Extract question data
                ques_json = {
                    "type": category_sheet.cell(row=i, column=1).value,
                    "pair1": category_sheet.cell(row=i, column=2).value,
                    "pair2": category_sheet.cell(row=i, column=3).value,
                    "img_link": category_sheet.cell(row=i, column=4).value,
                    "que": category_sheet.cell(row=i, column=5).value,
                    "opt1": category_sheet.cell(row=i, column=6).value,
                    "opt2": category_sheet.cell(row=i, column=7).value,
                    "opt3": category_sheet.cell(row=i, column=8).value,
                    "opt4": category_sheet.cell(row=i, column=9).value,
                    "correct_ans": category_sheet.cell(row=i, column=10).value,
                    "explanation": category_sheet.cell(row=i, column=11).value,
                    "level": category_sheet.cell(row=i, column=12).value,
                    "subcategory": category_sheet.cell(row=i, column=13).value,
                }

                # Stop if the row is empty
                if all(val is None for val in ques_json.values()):
                    break

                # Validate question
                comment = ""
                for validation_func in [
                    type_validation, pair_validation, options_validation,
                    explanation_validation, subcategory_validation, correct_ans_validation,
                    level_validation, IMG_validation, img_link_validation_MCQ_MPF,
                    pair_validation_MCQ_IMG
                ]:
                    valid, validation_comment = validation_func(ques_json)
                    if not valid:
                        comment += ";" + validation_comment

                if not comment:
                    comment = "Valid Question"
                    question_set_for_category.append(ques_json)
                    category_sheet.cell(row=i, column=14).value = "Valid Question"
                else:
                    validation_success = False
                    category_sheet.cell(row=i, column=14).value = comment[1:]
                    for k in range(1, 14):
                        category_sheet.cell(row=i, column=k).fill = redFill
                i += 1

            # Add questions to the DataFrame list
            df1 = pd.DataFrame.from_dict(question_set_for_category)
            df1['category_id'] = catId
            list_of_dfs.append(df1)

        # Handle subcategories
        for df in list_of_dfs:
            for _, row in df.iterrows():
                subcategory = row['subcategory']
                category_id = row['category_id']
                if subcategory not in subcategories:
                    subcategories[subcategory] = {'category_ids': []}
                if category_id not in subcategories[subcategory]['category_ids']:
                    subcategories[subcategory]['category_ids'].append(category_id)

        # Add subcategories to the database
        for subcategory, info in subcategories.items():
            for category_id in info['category_ids']:
                existing_subcategory = Subcategory.query.filter_by(
                    topic_name=subcategory, category_id=category_id, admin_id=adminID
                ).first()
                if not existing_subcategory:
                    new_subcategory = Subcategory(
                        topic_name=subcategory,
                        category_id=category_id,
                        admin_id=adminID
                    )
                    db.session.add(new_subcategory)

        db.session.commit()

        # Save results to an Excel file
        wb_obj.save("Validation Results.xlsx")
        matched_df = pd.concat(list_of_dfs, ignore_index=True)

        # Add questions to the quiz table
        for _, json_data in matched_df.iterrows():
            existing_quiz = Quize.query.filter_by(
                que=json_data['que'], category_id=json_data['category_id'],
                subcategory=json_data['subcategory']
            ).first()
            if not existing_quiz:
                new_quiz = Quize(
                    type=json_data['type'],
                    pair1=json_data['pair1'],
                    pair2=json_data['pair2'],
                    img_link=json_data['img_link'],
                    que=json_data['que'],
                    opt1=json_data['opt1'],
                    opt2=json_data['opt2'],
                    opt3=json_data['opt3'],
                    opt4=json_data['opt4'],
                    correct_ans=json_data['correct_ans'],
                    subcategory=json_data['subcategory'],
                    explanation=json_data['explanation'],
                    level=json_data['level'],
                    category_id=json_data['category_id'],
                    adminID=adminID
                )
                db.session.add(new_quiz)

        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        validation_success = False
        print(f"Database error: {e}")

    # Return response
    msg = {'msg': 'success' if validation_success else 'failed'}
    return jsonify(msg)







# Run Server
if __name__ =='__main__':
    app.run()

    db.session.query(Quize).delete()
    db.session.commit()    
    















         




                                                    








    

    

